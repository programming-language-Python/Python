import getpass
import os
from os.path import dirname, join
from openpyxl import load_workbook
# 1) Парсинг excel, занести в список
# 2) Папка на рабочем столе с номерами группы
# 3) Для каждого студента создав папку с его ФИО

# берёт текущий путь каталога
current_dir = dirname(__file__)
# к current_dir доавляем путь
file_path = join(current_dir, "Список группы.xlsx")
print(file_path)

wb = load_workbook(file_path)

# Получить лист по имени
sheet = wb['Лист1']
# Получить максимальное количество строк
count_row = sheet.max_row + 1
students = []
# Распечатать значения в столбце 2
for i in range(1, count_row):
    student = sheet.cell(row=i, column=2).value
    students.append(student)
# os.getenv("SystemDrive") - системный диск
# getpass.getuser() - имя пользователя
path = os.getenv("SystemDrive") + '\\Users\\' + \
    getpass.getuser() + '\\Desktop\\407'
for student in students:
    # создаёт каталог если такого каталога нет иначе выведет ошибку
    os.makedirs(path + '\\' + student)
